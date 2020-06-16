import os
import openpyxl
import shutil
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

alignment_all = Alignment(horizontal='center',
                          vertical='center',
                          wrap_text=False)
alignment_col1 = Alignment(horizontal='right',
                           vertical='center',
                           wrap_text=False)
font = Font(name='宋体')
thin = Side(border_style="thin", color="FF000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)

def mark_position(values): # for titles in row one
    """param 'values' is a list which contains 'None'
       as its elements, 'None' values can be continous.

       this function would mark the span of continous 'None' values
       and return a list"""
    mark = []
    a = False
    temp = []
    for each in values:
        if each == None and a == False:
            if not mark:
                temp.append(values.index(each))
                a = True
            else:
                temp.append(values[mark[-1][1]:].index(each) + mark[-1][1])
                a = True
        if each != None and a == True:
            temp.append(values.index(each))
            a = False
            mark.append(temp[:])
            temp.clear()
    mark = [[i+1 for i in each] for each in mark]
    return mark #需要在位置上向右位移一位

def beautify_sheet(sheet, columns, rows):
    cells = []
    col = []
    for i in range(1, columns + 1):
        column = getColumnLetter(i)
        col.append(column)
        sheet.column_dimensions[column].width = 30
    for i in range(1, rows + 1):
        sheet.row_dimensions[i].height = 30
        cells.extend([c + str(i) for c in col])
    for each in cells:
        sheet[each].font = font
        sheet[each].border = border
        if 'A' not in each:
            sheet[each].alignment = alignment_all
        else:
            sheet[each].alignment = alignment_col1

def produce_book(name, data):
    columns = len(data[0])
    rows = len(data)
    book = openpyxl.Workbook()
    sheet = book.worksheets[0]
    print('提取[%s]成功' % name)
    run = [sheet.append(i) for i in data]
    beautify_sheet(sheet, columns, rows)
    book.save(name)
    book.close()

def getColumnLetter(number):
    return openpyxl.cell.cell.get_column_letter(number)

def vertical_data(items, data):
    return list(zip(items, data))

def get_all_books():
    books = [_ for _ in os.listdir() if os.path.isfile(_) and '工资' in _]
    books = [os.path.join(os.getcwd(), _) for _ in books]
    return books

def extract(workbook):
    wb = openpyxl.open(workbook)
    ws = wb['三分局a']
    salary_info = [[i.value for i in each] for each in ws]
    title = salary_info[0]
    persons = salary_info[1:]
    row = 1
    for each in persons:
        try:
            name = each[1] + '_' + each[0] + '.xlsx'
            row += 1
        except TypeError as e:
            row += 1
            with open('logs.txt','a',encoding='utf-8') as f:
                info = str(each[1])+' _ '+ str(each[0])
                location = ' 第%s行 ##'% row
                content = workbook.split('\\')[-1] + '##' + location + info + ':' + str(e)+'\n'
                f.write(content)
            continue
        data = vertical_data(title, each)
        produce_book(name, data)
    wb.close()

def sort_files(suffix):
    print('\n整理文件中...\n')
    files = [_ for _ in os.listdir() if os.path.splitext(_)[1]=='.xlsx']
    names = [_.split('_')[0] for _ in files]
    names_unique = set(names)
    for name in names_unique:
        folder = name + suffix
        os.mkdir(folder)
        for file in files:
            if name == file.split('_')[0]:
                shutil.move(file, folder)
    print('\n整理完成！\n')

if __name__ == "__main__":
    books = get_all_books()
    if not os.path.exists('个人工资信息'):
        os.mkdir('个人工资信息')
        os.chdir('个人工资信息')
    else:
        os.chdir('个人工资信息')
    for book in books:
        extract(book)
    sort_files('_2019年工资情况')
