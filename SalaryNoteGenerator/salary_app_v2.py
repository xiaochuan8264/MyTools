print('\n\n\t等待5-30秒，小程序加载中...\n\n\n\t如有任何疑问，请不要问！！！\n\n')
import tkinter as tk
from openpyxl import open as openxl, Workbook
import tkinter.filedialog
import os
from time import sleep
import shutil
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.cell.cell import get_column_letter

alignment_all = Alignment(horizontal='center',
                          vertical='center',
                          wrap_text=False)
alignment_col1 = Alignment(horizontal='right',
                           vertical='center',
                           wrap_text=False)
font = Font(name='宋体')
thin = Side(border_style="thin", color="FF000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)

def openFile():
    def progress():
        tk.Lable(window, text='提取进度：',).place(x=50, y=60)
        canvas = tk.Canvas(window, width=300, height=22, bg="white")
        canvas.place(x=110, y=60)
        fill_line = canvas.create_rectangle(1.5, 1.5, 0, 23, width=0, fill="green")

    def operate_new(sheet_name):
        note = '%s 提取进度'%sheet_name
        label = tk.Label(window, text=note,)
        label.place(x=15, y=250)
        canvas = tk.Canvas(window, width=300, height=22, bg="white")
        canvas.place(x=150, y=250)
        fill_line = canvas.create_rectangle(1.5, 1.5, 0, 23, width=0, fill="green")
        ws = wb[sheet_name]
        salary_info = [[i.value for i in each] for each in ws]
        title = salary_info[0]
        persons = salary_info[1:]
        count = len(persons)
        n = 300/count # 300是矩形填充满的次数
        for each in persons:
            try:
                n += 300/count
                name = each[1] + '_' + each[0] + '.xlsx'
            except TypeError as e:
                with open('logs.txt','a',encoding='utf-8') as f:
                    info = str(each[1])+str(each[0])
                    f.write(info + ':' + str(e)+'\n')
                continue
            data = list(zip(title, each))
            produce_book(name, data)
            canvas.coords(fill_line, (0, 0, n, 250))
            window.update()
        sleep(0.5)
        label.destroy()
        canvas.destroy()

    def getSheet(event):
        choice = LB.get(LB.curselection())
        # print(type(choice))
        sheet = choice.split('"')[1]
        print(choice)
        print('\n开始处理表单...\n')
        operate_new(sheet)
        LB.destroy()

    f = tk.filedialog.askopenfilename(filetypes=[('Excel文件','.xlsx')])
    wb = openxl(f)
    sheets = wb.worksheets
    LB = tk.Listbox(window, width=50)
    a = LB.bind('<Double-Button-1>',getSheet)
    for i in sheets:
        LB.insert(tk.END, i)
    LB.pack()

def beautify_sheet(sheet, columns, rows):
    cells = []
    col = []
    for i in range(1, columns + 1):
        column = get_column_letter(i)
        col.append(column)
        sheet.column_dimensions[column].width = 30
    for i in range(1, rows + 1):
        sheet.row_dimensions[i].height = 30
        cells.extend([c + str(i) for c in col])
    for each in cells:
        sheet[each].font = font
        sheet[each].border = border
        if 'A' not in each:
            sheet[each].alignment = alignment_all
        else:
            sheet[each].alignment = alignment_col1

def produce_book(name, data):
    columns = len(data[0])
    rows = len(data)
    book = Workbook()
    sheet = book.worksheets[0]
    print('提取[%s]成功' % name)
    run = [sheet.append(i) for i in data]
    beautify_sheet(sheet, columns, rows)
    book.save(name)
    book.close()

def make_n_change_dir(name='个人工资信息'):
    if not os.path.exists(name):
        os.mkdir(name)
        os.chdir(name)
    else:
        os.chdir(name)

window = tk.Tk()
window.title('个人工资信息提取APP')
window.geometry('500x300')
choice = tk.Button(window,text='打开工作簿', command=openFile)
choice.pack()
make_n_change_dir()
window.mainloop()
